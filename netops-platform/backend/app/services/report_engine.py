"""
报表引擎：聚合平台数据并生成日报 PDF/Excel。
当前为最小可用版本，模板写死；后续可扩展为模板化配置。
"""
from __future__ import annotations

import os
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import httpx
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import func, select

from app.db import async_session
from app.models import Device, Alert
from app.config import settings

logger = logging.getLogger(__name__)

REPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)


class ReportData:
    """日报所需的数据聚合结果"""
    def __init__(
        self,
        generated_at: datetime,
        period_start: datetime,
        period_end: datetime,
        devices_total: int,
        devices_online: int,
        devices_offline: int,
        devices_error: int,
        active_alerts: int,
        alert_stats: dict,
        cpu_top5: list,
        memory_top5: list,
        traffic_top5: list,
    ):
        self.generated_at = generated_at
        self.period_start = period_start
        self.period_end = period_end
        self.devices_total = devices_total
        self.devices_online = devices_online
        self.devices_offline = devices_offline
        self.devices_error = devices_error
        self.active_alerts = active_alerts
        self.alert_stats = alert_stats
        self.cpu_top5 = cpu_top5
        self.memory_top5 = memory_top5
        self.traffic_top5 = traffic_top5


async def collect_daily_report_data(hours: int = 24) -> ReportData:
    """聚合日报数据"""
    now = datetime.utcnow()
    start = now - timedelta(hours=hours)

    async with async_session() as session:
        total = (await session.execute(select(func.count()).select_from(Device))).scalar() or 0
        online = (await session.execute(select(func.count()).where(Device.status == "online"))).scalar() or 0
        offline = (await session.execute(select(func.count()).where(Device.status == "offline"))).scalar() or 0
        error = (await session.execute(select(func.count()).where(Device.status == "error"))).scalar() or 0

        active_alerts = (await session.execute(
            select(func.count()).where(Alert.status == "firing")
        )).scalar() or 0

        # 按严重度统计告警
        severity_counts = {}
        for sev in ["critical", "warning", "info"]:
            cnt = (await session.execute(
                select(func.count()).where(Alert.status == "firing", Alert.severity == sev)
            )).scalar() or 0
            severity_counts[sev] = cnt

    cpu_top5 = await _query_top_metric("snmp_cpu_usage_percent", "avg", topn=5, hours=hours)
    memory_top5 = await _query_top_metric("snmp_memory_usage_percent", "avg", topn=5, hours=hours)
    traffic_top5 = await _query_top_traffic(topn=5, hours=hours)

    return ReportData(
        generated_at=now,
        period_start=start,
        period_end=now,
        devices_total=total,
        devices_online=online,
        devices_offline=offline,
        devices_error=error,
        active_alerts=active_alerts,
        alert_stats=severity_counts,
        cpu_top5=cpu_top5,
        memory_top5=memory_top5,
        traffic_top5=traffic_top5,
    )


async def _query_top_metric(metric_name: str, aggregation: str, topn: int, hours: int) -> list:
    """从 VictoriaMetrics 查询 Top N 指标"""
    try:
        end = datetime.utcnow()
        start = end - timedelta(hours=hours)
        start_str = start.isoformat() + "Z"
        end_str = end.isoformat() + "Z"

        # PromQL: avg_over_time / max_over_time
        if aggregation == "avg":
            promql = f"avg_over_time({metric_name}[{hours}h])"
        else:
            promql = f"max_over_time({metric_name}[{hours}h])"

        url = f"{settings.vm_url}/api/v1/query_range"
        params = {
            "query": promql,
            "start": start_str,
            "end": end_str,
            "step": "3600s",
        }
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(url, params=params)
            resp.raise_for_status()
            data = resp.json().get("data", {}).get("result", [])

        results = []
        for item in data:
            device_name = item.get("metric", {}).get("device", "unknown")
            values = item.get("values", [])
            if values:
                # 取最后一个值作为当前快照
                last_value = float(values[-1][1])
                results.append({"device": device_name, "value": round(last_value, 2)})

        results.sort(key=lambda x: x["value"], reverse=True)
        return results[:topn]
    except Exception as e:
        logger.error(f"查询 {metric_name} top{topn} 失败: {e}")
        return []


async def _query_top_traffic(topn: int, hours: int) -> list:
    """查询接口入向流量 Top N（bps）"""
    try:
        end = datetime.utcnow()
        start = end - timedelta(hours=hours)
        start_str = start.isoformat() + "Z"
        end_str = end.isoformat() + "Z"

        promql = f"max_over_time(snmp_interface_in_bits_per_second[{hours}h])"
        url = f"{settings.vm_url}/api/v1/query_range"
        params = {
            "query": promql,
            "start": start_str,
            "end": end_str,
            "step": "3600s",
        }
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(url, params=params)
            resp.raise_for_status()
            data = resp.json().get("data", {}).get("result", [])

        results = []
        for item in data:
            metric = item.get("metric", {})
            device_name = metric.get("device", "unknown")
            interface = metric.get("interface", "unknown")
            values = item.get("values", [])
            if values:
                last_value = float(values[-1][1])
                results.append({
                    "device": device_name,
                    "interface": interface,
                    "value_bps": round(last_value, 2),
                })

        results.sort(key=lambda x: x["value_bps"], reverse=True)
        return results[:topn]
    except Exception as e:
        logger.error(f"查询接口流量 top{topn} 失败: {e}")
        return []


async def generate_daily_pdf(data: ReportData) -> Path:
    """生成日报 PDF"""
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # 中文字体：系统没有 SimHei 等字体时fallback到内置Helvetica，会显示乱码。
    # 生产环境建议把 SimHei/SimSun 字体文件放到 backend/fonts/ 目录后注册。
    font_path = _find_chinese_font()
    if font_path:
        pdf.add_font("CustomCJK", "", str(font_path), uni=True)
        pdf.set_font("CustomCJK", "", 12)
    else:
        pdf.set_font("Helvetica", "", 12)

    # 标题
    pdf.set_font_size(18)
    pdf.cell(0, 12, "NetOps 网络运维日报", ln=True, align="C")
    pdf.set_font_size(10)
    pdf.cell(0, 8, f"生成时间: {data.generated_at.strftime('%Y-%m-%d %H:%M UTC')}", ln=True, align="C")
    pdf.ln(5)

    # 设备概览
    pdf.set_font_size(14)
    pdf.cell(0, 10, "一、设备运行概览", ln=True)
    pdf.set_font_size(11)
    summary = (
        f"设备总数: {data.devices_total}    "
        f"在线: {data.devices_online}    "
        f"离线: {data.devices_offline}    "
        f"异常: {data.devices_error}"
    )
    pdf.cell(0, 8, summary, ln=True)
    pdf.ln(3)

    # 告警概览
    pdf.set_font_size(14)
    pdf.cell(0, 10, "二、告警概览", ln=True)
    pdf.set_font_size(11)
    pdf.cell(0, 8, f"当前未恢复告警: {data.active_alerts}", ln=True)
    if data.alert_stats:
        sev_text = " / ".join(f"{k}: {v}" for k, v in data.alert_stats.items())
        pdf.cell(0, 8, f"按严重度分布: {sev_text}", ln=True)
    pdf.ln(3)

    # TOP5 CPU
    _write_table(pdf, "三、CPU 使用率 TOP5", ["设备", "平均 CPU%"], data.cpu_top5, "device", "value")

    # TOP5 Memory
    _write_table(pdf, "四、内存使用率 TOP5", ["设备", "平均内存%"], data.memory_top5, "device", "value")

    # TOP5 Traffic
    _write_table(pdf, "五、接口入向流量 TOP5", ["设备", "接口", "峰值(bps)"], data.traffic_top5, "device", "interface", "value_bps")

    # 结论
    pdf.set_font_size(14)
    pdf.cell(0, 10, "六、巡检结论", ln=True)
    pdf.set_font_size(11)
    if data.devices_offline > 0 or data.devices_error > 0:
        conclusion = f"本周期内共有 {data.devices_offline} 台设备离线、{data.devices_error} 台设备异常，{data.active_alerts} 条未恢复告警，建议优先处理离线/异常设备及 critical 告警。"
    else:
        conclusion = f"本周期内网络整体运行平稳，{data.devices_online}/{data.devices_total} 台设备在线，当前无未恢复告警。"
    pdf.multi_cell(0, 8, conclusion)

    output_path = REPORTS_DIR / f"daily_report_{data.generated_at.strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(str(output_path))
    return output_path


def _write_table(pdf: FPDF, title: str, headers: list, rows: list, *keys: str):
    pdf.set_font_size(14)
    pdf.cell(0, 10, title, ln=True)
    pdf.set_font_size(10)

    if not rows:
        pdf.cell(0, 8, "（暂无数据）", ln=True)
        pdf.ln(3)
        return

    # 表头背景
    pdf.set_fill_color(230, 230, 230)
    col_widths = [170 / len(headers)] * len(headers)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, str(h), border=1, fill=True, align="C")
    pdf.ln()

    # 行
    for row in rows:
        vals = [str(row.get(k, "-")) for k in keys]
        for i, v in enumerate(vals):
            pdf.cell(col_widths[i], 8, v, border=1, align="C")
        pdf.ln()
    pdf.ln(3)


async def generate_daily_excel(data: ReportData) -> Path:
    """生成日报 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "日报"

    # 样式
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=12, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"] = "NetOps 网络运维日报"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"生成时间: {data.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
    ws.merge_cells("A2:F2")

    row = 4
    ws[f"A{row}"] = "一、设备运行概览"
    ws[f"A{row}"].font = header_font
    row += 1
    ws.append(["设备总数", "在线", "离线", "异常", "在线率"])
    online_rate = f"{data.devices_online / data.devices_total * 100:.1f}%" if data.devices_total else "0%"
    ws.append([data.devices_total, data.devices_online, data.devices_offline, data.devices_error, online_rate])

    row = ws.max_row + 2
    ws[f"A{row}"] = "二、告警概览"
    ws[f"A{row}"].font = header_font
    row += 1
    ws.append(["未恢复告警", "critical", "warning", "info"])
    ws.append([data.active_alerts, data.alert_stats.get("critical", 0), data.alert_stats.get("warning", 0), data.alert_stats.get("info", 0)])

    # TOP5 CPU
    row = ws.max_row + 2
    ws[f"A{row}"] = "三、CPU 使用率 TOP5"
    ws[f"A{row}"].font = header_font
    ws.append(["设备", "平均 CPU%"])
    for item in data.cpu_top5:
        ws.append([item["device"], item["value"]])

    # TOP5 Memory
    row = ws.max_row + 2
    ws[f"A{row}"] = "四、内存使用率 TOP5"
    ws[f"A{row}"].font = header_font
    ws.append(["设备", "平均内存%"])
    for item in data.memory_top5:
        ws.append([item["device"], item["value"]])

    # TOP5 Traffic
    row = ws.max_row + 2
    ws[f"A{row}"] = "五、接口入向流量 TOP5"
    ws[f"A{row}"].font = header_font
    ws.append(["设备", "接口", "峰值(bps)"])
    for item in data.traffic_top5:
        ws.append([item["device"], item["interface"], item["value_bps"]])

    # 结论
    row = ws.max_row + 2
    ws[f"A{row}"] = "六、巡检结论"
    ws[f"A{row}"].font = header_font
    row += 1
    if data.devices_offline > 0 or data.devices_error > 0:
        conclusion = f"本周期内共有 {data.devices_offline} 台设备离线、{data.devices_error} 台设备异常，{data.active_alerts} 条未恢复告警，建议优先处理离线/异常设备及 critical 告警。"
    else:
        conclusion = f"本周期内网络整体运行平稳，{data.devices_online}/{data.devices_total} 台设备在线，当前无未恢复告警。"
    ws[f"A{row}"] = conclusion
    ws.merge_cells(f"A{row}:F{row}")

    output_path = REPORTS_DIR / f"daily_report_{data.generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(output_path))
    return output_path


def _find_chinese_font() -> Optional[Path]:
    """寻找可用的中文字体文件"""
    candidates = [
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
        Path("C:/Windows/Fonts/msyh.ttc"),
    ]
    fonts_dir = Path(__file__).resolve().parent.parent.parent / "fonts"
    if fonts_dir.exists():
        for f in fonts_dir.iterdir():
            if f.suffix.lower() in (".ttf", ".ttc"):
                candidates.append(f)
    for p in candidates:
        if p.exists():
            return p
    return None
